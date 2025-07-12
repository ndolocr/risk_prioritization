import io
import xlwt
import urllib
import base64
import random
import openpyxl
import itertools
import numpy as np
import skfuzzy as fuzz
import matplotlib
matplotlib.use('Agg')
from datetime import datetime
import matplotlib.pyplot as plt
from skfuzzy import control as ctrl

from django.shortcuts import render
from django.http import HttpResponse
from django.utils.text import slugify

from openpyxl.utils import get_column_letter

# from core.utils import rules_with_cost
from core.utils import fuzzy_engine

#Calling function from fuzzy_engine file to get geerated rules, risk score
#and antecedents for DREAD and DREAD-C
control_system_with_cost, antecedents_with_cost, risk_score_cost = fuzzy_engine.get_fuzzy_risk_control_system_with_cost_and_dread()
control_system_without_cost, antecedents_without_cost, risk_score = fuzzy_engine.get_fuzzy_risk_control_system_without_cost()
# control_system_with_cost = 10
# antecedents_with_cost = 10
# risk_score_cost = 10
# control_system_without_cost = 10
# antecedents_without_cost = 10 
# risk_score = 10

# Calculate risk for DREAD and COST.
def risk_with_cost_for_windows(request):
    if request.method == 'POST':

        # Get request from browser
        cost_input = float(request.POST.get('cost'))
        exploitability_input = float(request.POST.get('exploitability'))
        affected_users_input = float(request.POST.get('affected_users'))
        discoverability_input = float(request.POST.get('discoverability'))
        reproducibility_input = float(request.POST.get('reproducibility'))
        damage_potential_input = float(request.POST.get('damage_potential'))

        #Loading rules for DREAD nd DREAD-C
        dreadc_sim = ctrl.ControlSystemSimulation(control_system_with_cost)
        dread_sim = ctrl.ControlSystemSimulation(control_system_without_cost)

        #Loading input values from brower an rules generated for DREAD
        dread_sim.input['exploitability'] = exploitability_input
        dread_sim.input['affected_users'] = affected_users_input
        dread_sim.input['discoverability'] = discoverability_input
        dread_sim.input['reproducibility'] = reproducibility_input
        dread_sim.input['damage_potential'] = damage_potential_input

        #Compute score fr DREAD
        dread_sim.compute()
        dread_sim_result = dread_sim.output['risk_score']
        print(f"DREAD -- > {dread_sim_result}")
        print(f"COST -- > {cost_input}")
        #Loading input values from brower an rules generated for DREAD-C
        dreadc_sim.input['cost'] = cost_input
        dreadc_sim.input['dread'] = dread_sim_result
        
        #Compute score fr DREAD-C
        dreadc_sim.compute()
        dreadc_sim_result = dreadc_sim.output['risk_score']

        print(f"DREAD - C --> {dreadc_sim_result}")

        # Plot risk score output
        plt.figure()
        risk_score_cost.view(sim=dreadc_sim)
        buf = io.BytesIO()
        plt.savefig(buf, format='png')
        buf.seek(0)
        graph_1 = base64.b64encode(buf.read()).decode('utf-8')
        buf.close()

        # Plot a simple bar chart showing individual and combined scores
        labels = ['Dread Risk', 'Dread-C Risk']
        values = [dread_sim_result, dreadc_sim_result]

        plt.figure(figsize=(8, 5))
        bars = plt.bar(labels, values, color=['#ff9999','#66b3ff'])
        # plt.ylim(0, 110)
        plt.ylim(0, 11)  
        plt.title('Fuzzy Risk Evaluation')
        plt.ylabel('Risk Score')

        # Annotate bars
        for bar in bars:
            height = bar.get_height()
            plt.text(bar.get_x() + bar.get_width() / 2., height + 0.5,
                    f'{height:.2f}', ha='center', va='bottom')

        buf = io.BytesIO()
        plt.tight_layout()
        plt.savefig(buf, format='png')
        buf.seek(0)
        graph = base64.b64encode(buf.read()).decode('utf-8')
        buf.close()

        #Generate output data to the browser 
        input_dict = {
            'cost': cost_input,
            'result_cost_sim': round(dreadc_sim_result, 2),
            'result_risk_sim': round(dread_sim_result, 2),
            'exploitability': exploitability_input, 
            'affected_users': affected_users_input, 
            'discoverability': discoverability_input, 
            'reproducibility': reproducibility_input, 
            'damage_potential': damage_potential_input
            }
        
        #Generating Decision
        decision=""

        if dreadc_sim_result <= 3:
            decision = "Accept Risk."
        elif dreadc_sim_result > 3 and dreadc_sim_result < 4:
            decision = "Transfer Risk."
        elif dreadc_sim_result >= 4 and dreadc_sim_result <= 5:
            decision = "Reduce Risk."
        elif dreadc_sim_result > 5:
            decision = "Avoid Risk."

        # if dreadc_sim_result <= 30:
        #     decision = "Accept Risk."
        # elif dreadc_sim_result > 30 and dreadc_sim_result < 40:
        #     decision = "Transfer Risk."
        # elif dreadc_sim_result >= 40 and dreadc_sim_result <= 50:
        #     decision = "Reduce Risk."
        # elif dreadc_sim_result > 50:
        #     decision = "Avoid Risk."
        
        #Creating object to be rendered on browser.
        context = {
            'graph': graph, 
            'graph_1': graph_1,            
            'input': input_dict,
            'decision': decision,
            'result': round(dreadc_sim_result, 2),
        }

        #return html page t show resuts, together ith object and data to b rendered,
        return render(request, 'core/risk_with_cost_windows.html', context)
    
    return render(request, 'core/risk_with_cost_windows.html')

def risk_without_cost_for_windows(request):
    if request.method == 'POST':        
        exploitability_input = int(request.POST.get('exploitability'))
        affected_users_input = int(request.POST.get('affected_users'))
        discoverability_input = int(request.POST.get('discoverability'))
        reproducibility_input = int(request.POST.get('reproducibility'))
        damage_potential_input = int(request.POST.get('damage_potential'))

        cost_sim = ctrl.ControlSystemSimulation(control_system_with_cost)
        risk_sim = ctrl.ControlSystemSimulation(control_system_without_cost)

        risk_sim.input['exploitability'] = exploitability_input
        risk_sim.input['affected_users'] = affected_users_input
        risk_sim.input['discoverability'] = discoverability_input
        risk_sim.input['reproducibility'] = reproducibility_input
        risk_sim.input['damage_potential'] = damage_potential_input

        risk_sim.compute()
        
        result_risk_sim =risk_sim.output['risk_score']

        result = result_risk_sim

        # Plot risk score output
        plt.figure()
        risk_score.view(sim=risk_sim)
        buf = io.BytesIO()
        plt.savefig(buf, format='png')
        buf.seek(0)
        graph_1 = base64.b64encode(buf.read()).decode('utf-8')
        buf.close()

        # Plot a simple bar chart showing individual and combined scores
        labels = ['Threat Risk']
        values = [result]

        plt.figure(figsize=(8, 5))
        bars = plt.bar(labels, values, color=['#ff9999'])
        plt.ylim(0, 110)  # max score is 10 + 10
        plt.title('Fuzzy Risk Evaluation')
        plt.ylabel('Risk Score')

        # Annotate bars
        for bar in bars:
            height = bar.get_height()
            plt.text(bar.get_x() + bar.get_width() / 2., height + 0.5,
                    f'{height:.2f}', ha='center', va='bottom')

        buf = io.BytesIO()
        plt.tight_layout()
        plt.savefig(buf, format='png')
        buf.seek(0)
        graph = base64.b64encode(buf.read()).decode('utf-8')
        buf.close()

        input_dict = {
            'result_risk_sim': round(result_risk_sim, 2),
            'exploitability': exploitability_input, 
            'affected_users': affected_users_input, 
            'discoverability': discoverability_input, 
            'reproducibility': reproducibility_input, 
            'damage_potential': damage_potential_input
            }
        
        context = {
            'graph': graph, 
            'graph_1': graph_1,            
            'input': input_dict,
            'result': round(result, 2),
        }

        return render(request, 'core/risk_without_cost_windows.html', context)
    
    return render(request, 'core/risk_without_cost_windows.html')











def risk_with_cost_for_macbook(request):
    if request.method == 'POST':

        cost_input = int(request.POST.get('cost'))
        exploitability_input = int(request.POST.get('exploitability'))
        affected_users_input = int(request.POST.get('affected_users'))
        discoverability_input = int(request.POST.get('discoverability'))
        reproducibility_input = int(request.POST.get('reproducibility'))
        damage_potential_input = int(request.POST.get('damage_potential'))

        cost_sim = ctrl.ControlSystemSimulation(control_system_with_cost)
        risk_sim = ctrl.ControlSystemSimulation(control_system_without_cost)
        
        cost_sim.input['cost'] = cost_input

        risk_sim.input['exploitability'] = exploitability_input
        risk_sim.input['affected_users'] = affected_users_input
        risk_sim.input['discoverability'] = discoverability_input
        risk_sim.input['reproducibility'] = reproducibility_input
        risk_sim.input['damage_potential'] = damage_potential_input

        risk_sim.compute()
        cost_sim.compute()
        
        result_cost_sim =cost_sim.output['cost_score']
        result_risk_sim =risk_sim.output['risk_score']

        print(f"Cost Score --> {result_cost_sim}")
        print(f"Risk Score --> {result_risk_sim}")

        # result = risk_sim.output['risk_score']
        result = result_cost_sim + result_risk_sim

        print(f"Total Risk Score --> {result}")

        # Plot
        fig, ax = plt.subplots(figsize=(10, 6))

        # Bar chart for risk score
        ax.bar(['Risk Score'], [result], color='orange', width=0.4, label='Fuzzy Risk Score')

        # Overlay line chart for inputs
        input_labels = ['cost', 'Exploitability', 'Affected Users', 'Discoverability', 'Reproducibility', 'Damage Potential']

        input_values = [
            cost_input,
            exploitability_input, 
            affected_users_input, 
            discoverability_input, 
            reproducibility_input, 
            damage_potential_input
        ]

        ax.plot(input_labels, input_values, color='blue', marker='o', linestyle='-', label='Input Factors')

        ax.set_ylim(0, 15)
        ax.set_ylabel('Score (0-15)')
        ax.set_title('Fuzzy Risk Score with Input Factors')
        ax.legend()

        # Save and encode image
        buffer = io.BytesIO()
        plt.tight_layout()
        plt.savefig(buffer, format='png')
        buffer.seek(0)
        image_png = buffer.getvalue()
        buffer.close()

        graph = base64.b64encode(image_png).decode('utf-8')

        input_dict = {
            'cost': cost_input,
            'result_cost_sim': round(result_cost_sim, 2),
            'result_risk_sim': round(result_risk_sim, 2),
            'exploitability': exploitability_input,
            'affected_users': affected_users_input,
            'discoverability': discoverability_input,
            'reproducibility': reproducibility_input,
            'damage_potential': damage_potential_input
        }

        print(input_dict)
        
        # Pass risk score and graph to template
        context = {
            'graph': graph,
            'result': round(result, 2),
            'input': input_dict,
        }
        return render(request, 'core/risk_with_cost_for_macbook.html', context)
    return render(request, 'core/risk_with_cost_for_macbook.html')

def risk_without_cost_for_macbook(request):
    if request.method == 'POST':

        exploitability_input = int(request.POST.get('exploitability'))
        affected_users_input = int(request.POST.get('affected_users'))
        discoverability_input = int(request.POST.get('discoverability'))
        reproducibility_input = int(request.POST.get('reproducibility'))
        damage_potential_input = int(request.POST.get('damage_potential'))

        risk_sim = ctrl.ControlSystemSimulation(control_system_without_cost)

        risk_sim.input['exploitability'] = exploitability_input
        risk_sim.input['affected_users'] = affected_users_input
        risk_sim.input['discoverability'] = discoverability_input
        risk_sim.input['reproducibility'] = reproducibility_input
        risk_sim.input['damage_potential'] = damage_potential_input

        risk_sim.compute()
        result = risk_sim.output['risk_score']


        # Plot
        fig, ax = plt.subplots(figsize=(10, 6))

        # Bar chart for risk score
        ax.bar(['Risk Score'], [result], color='orange', width=0.4, label='Fuzzy Risk Score')

        # Overlay line chart for inputs
        input_labels = ['Exploitability', 'Affected Users', 'Discoverability', 'Reproducibility', 'Damage Potential']

        input_values = [
            exploitability_input, 
            affected_users_input, 
            discoverability_input, 
            reproducibility_input, 
            damage_potential_input
        ]

        ax.plot(input_labels, input_values, color='blue', marker='o', linestyle='-', label='Input Factors')

        ax.set_ylim(0, 15)
        ax.set_ylabel('Score (0-15)')
        ax.set_title('Fuzzy Risk Score with Input Factors')
        ax.legend()

        # Save and encode image
        buffer = io.BytesIO()
        plt.tight_layout()
        plt.savefig(buffer, format='png')
        buffer.seek(0)
        image_png = buffer.getvalue()
        buffer.close()

        graph = base64.b64encode(image_png).decode('utf-8')

        input_dict = {
            'exploitability': exploitability_input,
            'affected_users': affected_users_input,
            'discoverability': discoverability_input,
            'reproducibility': reproducibility_input,
            'damage_potential': damage_potential_input
        }

        print(input_dict)
        
        # Pass risk score and graph to template
        context = {
            'graph': graph,
            'result': round(result, 2),
            'input': input_dict,
        }
        return render(request, 'core/risk_without_cost_for_macbook.html', context)

    return render(request, 'core/risk_without_cost_for_macbook.html')

def generate_rules_with_cost():
    # Define membership levels
    levels = ['low', 'medium', 'high']
    antecedents = ['cost', 'damage_potential', 'exploitability', 'reproducibility', 'affected_users', 'discoverability']

    # Generate all 243 combinations
    combinations = list(itertools.product(levels, repeat=len(antecedents)))

    # Create rules list
    rules = []

    rules_number = 0
    for combo in combinations:
        rules_number = rules_number + 1
        condition = ' & '.join(f"{var}['{level}']" for var, level in zip(antecedents, combo))

        # Fixed fuzzy rule assignment logic
        if combo.count('high') >= 3:
            risk = "risk_score['high']"
        elif combo.count('low') >= 3:
            risk = "risk_score['low']"
        else:
            risk = "risk_score['medium']"

        rules.append(f"ctrl.Rule({condition}, {risk})")

    return [rules, rules_number]

def generate_rules_for_cost_only():
    # Define membership levels
    levels = ['low', 'medium', 'high']
    antecedents = ['cost']

    # Generate all 243 combinations
    combinations = list(itertools.product(levels, repeat=len(antecedents)))

    # Create rules list
    rules = []

    rules_number = 0
    for combo in combinations:
        rules_number = rules_number + 1
        condition = ' & '.join(f"{var}['{level}']" for var, level in zip(antecedents, combo))

        # Fixed fuzzy rule assignment logic
        if combo.count('high') >= 3:
            risk = "risk_score['high']"
        elif combo.count('low') >= 3:
            risk = "risk_score['low']"
        else:
            risk = "risk_score['medium']"

        rules.append(f"ctrl.Rule({condition}, {risk})")

    return [rules, rules_number]

def generate_rules_without_cost():
    # Define membership levels
    levels = ['low', 'medium', 'high']
    antecedents = ['damage_potential', 'exploitability', 'reproducibility', 'affected_users', 'discoverability']

    # Generate all 243 combinations
    combinations = list(itertools.product(levels, repeat=len(antecedents)))

    # Create rules list
    rules = []

    rules_number = 0
    for combo in combinations:
        rules_number = rules_number + 1
        condition = ' & '.join(f"{var}['{level}']" for var, level in zip(antecedents, combo))

        # Fixed fuzzy rule assignment logic
        if combo.count('high') >= 3:
            risk = "risk_score['high']"
        elif combo.count('low') >= 3:
            risk = "risk_score['low']"
        else:
            risk = "risk_score['medium']"

        rules.append(f"ctrl.Rule({condition}, {risk})")

    return [rules, rules_number]

def download_fuzzy_rules_with_cost(request):
    rules_response = generate_rules_with_cost()
    all_rules = rules_response[0]
    number_of_rules = rules_response[1]
    
    # Prepare text file content
    filename = f"fuzzy_rules_with_cost_{slugify(datetime.now().isoformat())}.txt"
    file_content = f"{number_of_rules} Rules\n\n" + "\n".join(all_rules)

    # Create response to download file
    response = HttpResponse(file_content, content_type='text/plain')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response

def download_fuzzy_rules_for_cost_only(request):
    rules_response = generate_rules_for_cost_only()
    all_rules = rules_response[0]
    number_of_rules = rules_response[1]
    
    # Prepare text file content
    filename = f"fuzzy_rules_cost_only_{slugify(datetime.now().isoformat())}.txt"
    file_content = f"{number_of_rules} Rules\n\n" + "\n".join(all_rules)

    # Create response to download file
    response = HttpResponse(file_content, content_type='text/plain')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response

def download_fuzzy_rules_without_cost(request):
    rules_response = generate_rules_without_cost()
    all_rules = rules_response[0]
    number_of_rules = rules_response[1]
    
    # Prepare text file content
    filename = f"fuzzy_rules_without_cost_{slugify(datetime.now().isoformat())}.txt"
    file_content = f"{number_of_rules} Rules\n\n" + "\n".join(all_rules)

    # Create response to download file
    response = HttpResponse(file_content, content_type='text/plain')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response

def view_fuzzy_rules_with_cost(request):
    rules_response = generate_rules_with_cost()
    all_rules = rules_response[0]
    number_of_rules = rules_response[1]
    print(f"{type(all_rules)}")
    # for rule in all_rules:
    #     print(f"{type(rule)}")
    context = {
        "all_rules": all_rules,
        "number_of_rules": number_of_rules,
    }

    return render(request, 'core/rules_with_cost.html', context)

def view_fuzzy_rules_without_cost(request):
    rules_response = generate_rules_without_cost()
    all_rules = rules_response[0]
    number_of_rules = rules_response[1]
    print(f"{type(all_rules)}")
    # for rule in all_rules:
    #     print(f"{type(rule)}")
    context = {
        "all_rules": all_rules,
        "number_of_rules": number_of_rules,
    }

    return render(request, 'core/rules_without_cost.html', context)

def view_fuzzy_rules_for_cost_only(request):
    rules_response = generate_rules_for_cost_only()
    all_rules = rules_response[0]
    number_of_rules = rules_response[1]
    print(f"{type(all_rules)}")
    # for rule in all_rules:
    #     print(f"{type(rule)}")
    context = {
        "all_rules": all_rules,
        "number_of_rules": number_of_rules,
    }

    return render(request, 'core/cost_rules_only.html', context)

#################################################################
def generate_dream_c_rules():
    # Define membership levels
    levels = ['low', 'medium', 'high']
    antecedents = ['cost', 'dread']

    # Generate all combinations
    combinations = list(itertools.product(levels, repeat=len(antecedents)))

    # Create rules list
    rules = []

    rules_number = 0
    for combo in combinations:
        rules_number = rules_number + 1
        condition = ' & '.join(f"{var}['{level}']" for var, level in zip(antecedents, combo))

        # Fixed fuzzy rule assignment logic
        if combo.count('high') >= 2:
            risk = "risk_score['high']"
        elif combo.count('low') >= 2:
            risk = "risk_score['low']"
        else:
            risk = "risk_score['medium']"

        rules.append(f"ctrl.Rule({condition}, {risk})")

    return [rules, rules_number]

def view_fuzzy_rules_for_dream_c(request):
    rules_response = generate_dream_c_rules()
    all_rules = rules_response[0]
    number_of_rules = rules_response[1]
    print(f"{type(all_rules)}")
 
    context = {
        "all_rules": all_rules,
        "number_of_rules": number_of_rules,
    }

    return render(request, 'core/dreamc_rules.html', context)

def download_dream_c_fuzzy_rules(request):
    rules_response = generate_dream_c_rules()
    all_rules = rules_response[0]
    number_of_rules = rules_response[1]
    
    # Prepare text file content
    filename = f"dream_c_fuzzy_rules_{slugify(datetime.now().isoformat())}.txt"
    file_content = f"{number_of_rules} Rules\n\n" + "\n".join(all_rules)

    # Create response to download file
    response = HttpResponse(file_content, content_type='text/plain')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response



# Generate risk with list values
def risk_with_cost_from_generated_list(request):
    list_size = 11

    final_list = []

    for val in range(0, list_size):
        cost_input = val
        exploitability_input = val
        affected_users_input = val
        discoverability_input = val
        reproducibility_input = val
        damage_potential_input = val

        dreadc_sim = ctrl.ControlSystemSimulation(control_system_with_cost)
        dread_sim = ctrl.ControlSystemSimulation(control_system_without_cost)

        #Loading input values from brower an rules generated for DREAD
        dread_sim.input['exploitability'] = exploitability_input
        dread_sim.input['affected_users'] = affected_users_input
        dread_sim.input['discoverability'] = discoverability_input
        dread_sim.input['reproducibility'] = reproducibility_input
        dread_sim.input['damage_potential'] = damage_potential_input

        #Compute score fr DREAD
        dread_sim.compute()
        dread_sim_result = dread_sim.output['risk_score']
        
        #Loading input values from brower an rules generated for DREAD-C
        dreadc_sim.input['cost'] = cost_input
        dreadc_sim.input['dread'] = dread_sim_result
        
        #Compute score fr DREAD-C
        dreadc_sim.compute()
        dreadc_sim_result = dreadc_sim.output['risk_score']

        decision=""

        if dreadc_sim_result <= 3:
            decision = "Accept Risk."
        elif dreadc_sim_result > 3 and dreadc_sim_result < 4:
            decision = "Transfer Risk."
        elif dreadc_sim_result >= 4 and dreadc_sim_result <= 5:
            decision = "Reduce Risk."
        elif dreadc_sim_result > 5:
            decision = "Avoid Risk."

        print("==================================")
        print(f"COST -- > {cost_input}")
        print(f"EXPLOITABILITY -- > {exploitability_input}")
        print(f"AFFECTED USERS -- > {affected_users_input}")
        print(f"DISCOVERABILITY -- > {discoverability_input}")
        print(f"REPRODUCIBILITY -- > {reproducibility_input}")
        print(f"DAMAGE CONTROL -- > {damage_potential_input}")
        print("------------------------------------")
        print(f"DREAD SCORE -- > {dread_sim_result}")
        print(f"DREAD - C SCORE --> {dreadc_sim_result}")

        obj = {
            'cost':cost_input,
            'decision': decision,            
            'exploitability':exploitability_input,
            'affected_users':affected_users_input,
            'discoverability':discoverability_input,
            'reproducibility':reproducibility_input,
            'damage_potential':damage_potential_input,
            'dread_score': round(dread_sim_result, 2),
            'dread_c_score': round(dreadc_sim_result, 2),
        }

        final_list.append(obj)
    context = {        
        "response": final_list,
        "title": "Incrementing Uniform Parameters",
    }

    return render(request, 'core/risk_with_cost_list_generated.html', context=context)
        
        
# Generate risk with list values
def risk_with_cost_from_generated_list_with_ascending_cost(request):
    list_size = 11
    random_number = round(random.uniform(0, 10), 2)
    final_list = []

    for val in range(0, list_size):
        cost_input = val
        exploitability_input = random_number
        affected_users_input = random_number
        discoverability_input = random_number
        reproducibility_input = random_number
        damage_potential_input = random_number

        dreadc_sim = ctrl.ControlSystemSimulation(control_system_with_cost)
        dread_sim = ctrl.ControlSystemSimulation(control_system_without_cost)

        #Loading input values from brower an rules generated for DREAD
        dread_sim.input['exploitability'] = exploitability_input
        dread_sim.input['affected_users'] = affected_users_input
        dread_sim.input['discoverability'] = discoverability_input
        dread_sim.input['reproducibility'] = reproducibility_input
        dread_sim.input['damage_potential'] = damage_potential_input

        #Compute score fr DREAD
        dread_sim.compute()
        dread_sim_result = dread_sim.output['risk_score']
        
        #Loading input values from brower an rules generated for DREAD-C
        dreadc_sim.input['cost'] = cost_input
        dreadc_sim.input['dread'] = dread_sim_result
        
        #Compute score fr DREAD-C
        dreadc_sim.compute()
        dreadc_sim_result = dreadc_sim.output['risk_score']

        decision=""

        if dreadc_sim_result <= 3:
            decision = "Accept Risk."
        elif dreadc_sim_result > 3 and dreadc_sim_result < 4:
            decision = "Transfer Risk."
        elif dreadc_sim_result >= 4 and dreadc_sim_result <= 5:
            decision = "Reduce Risk."
        elif dreadc_sim_result > 5:
            decision = "Avoid Risk."

        print("==================================")
        print(f"COST -- > {cost_input}")
        print(f"EXPLOITABILITY -- > {exploitability_input}")
        print(f"AFFECTED USERS -- > {affected_users_input}")
        print(f"DISCOVERABILITY -- > {discoverability_input}")
        print(f"REPRODUCIBILITY -- > {reproducibility_input}")
        print(f"DAMAGE CONTROL -- > {damage_potential_input}")
        print("------------------------------------")
        print(f"DREAD SCORE -- > {dread_sim_result}")
        print(f"DREAD - C SCORE --> {dreadc_sim_result}")

        obj = {
            'cost':cost_input,
            'decision': decision,            
            'exploitability':exploitability_input,
            'affected_users':affected_users_input,
            'discoverability':discoverability_input,
            'reproducibility':reproducibility_input,
            'damage_potential':damage_potential_input,
            'dread_score': round(dread_sim_result, 2),
            'dread_c_score': round(dreadc_sim_result, 2),
        }

        final_list.append(obj)
    context = {        
        "response": final_list,
        "title": "Incrementing Cost Parameters with constant DREAD Parameter",
    }

    return render(request, 'core/risk_with_cost_list_generated.html', context=context)

def risk_with_varying_dread_and_constant_cost_parameters(request):
    if request.method == "POST":
        cost = float(request.POST.get('cost', 0))
        list_number = int(request.POST.get('dread', 1000))

        numbers = []
        for _ in range(list_number):
            number = round(random.uniform(0, 10), 2)
            numbers.append(number)

        numbers.sort()

        final_list = []

        for val in numbers:
            cost_input = cost
            exploitability_input = val
            affected_users_input = val
            discoverability_input = val
            reproducibility_input = val
            damage_potential_input = val

            dreadc_sim = ctrl.ControlSystemSimulation(control_system_with_cost)
            dread_sim = ctrl.ControlSystemSimulation(control_system_without_cost)

            #Loading input values from brower an rules generated for DREAD
            dread_sim.input['exploitability'] = exploitability_input
            dread_sim.input['affected_users'] = affected_users_input
            dread_sim.input['discoverability'] = discoverability_input
            dread_sim.input['reproducibility'] = reproducibility_input
            dread_sim.input['damage_potential'] = damage_potential_input

            #Compute score fr DREAD
            dread_sim.compute()
            dread_sim_result = dread_sim.output['risk_score']
            
            #Loading input values from brower an rules generated for DREAD-C
            dreadc_sim.input['cost'] = cost_input
            dreadc_sim.input['dread'] = dread_sim_result
            
            #Compute score fr DREAD-C
            dreadc_sim.compute()
            dreadc_sim_result = dreadc_sim.output['risk_score']

            dread_sim_result = round(dread_sim_result, 2)
            dreadc_sim_result = round(dreadc_sim_result, 2)

            decision=""

            if dreadc_sim_result <= 3:
                decision = "Accept Risk."
            elif dreadc_sim_result > 3 and dreadc_sim_result < 4:
                decision = "Transfer Risk."
            elif dreadc_sim_result >= 4 and dreadc_sim_result <= 5:
                decision = "Reduce Risk."
            elif dreadc_sim_result > 5:
                decision = "Avoid Risk."

            print("==================================")
            print(f"COST -- > {cost_input}")
            print(f"EXPLOITABILITY -- > {exploitability_input}")
            print(f"AFFECTED USERS -- > {affected_users_input}")
            print(f"DISCOVERABILITY -- > {discoverability_input}")
            print(f"REPRODUCIBILITY -- > {reproducibility_input}")
            print(f"DAMAGE CONTROL -- > {damage_potential_input}")
            print("------------------------------------")
            print(f"DREAD SCORE -- > {dread_sim_result}")
            print(f"DREAD - C SCORE --> {dreadc_sim_result}")

            obj = {
                'cost':cost_input,
                'decision': decision,            
                'exploitability':exploitability_input,
                'affected_users':affected_users_input,
                'discoverability':discoverability_input,
                'reproducibility':reproducibility_input,
                'damage_potential':damage_potential_input,
                'dread_score': dread_sim_result,
                'dread_c_score': dreadc_sim_result,
            }

            final_list.append(obj)
        context = {        
            "response": final_list,
            "title": "Incrementing Cost Parameters with constant DREAD Parameter",
        }

        
        return render(request, 'core/risk_with_varying_dread_and_constant_cost_parameters.html', context=context)
    return render(request, 'core/risk_with_varying_dread_and_constant_cost_parameters.html')

def risk_with_varying_cost_and_constant_dread_parameters(request):
    if request.method == "POST":
        generate_nums = int(request.POST.get('generate_nums', 1))
        dread = float(request.POST.get('dread', 10))

        numbers = []
        for _ in range(generate_nums):
            number = round(random.uniform(0, 10), 2)
            numbers.append(number)

        numbers.sort()

        final_list = []

        for val in numbers:
            cost_input = val
            exploitability_input = dread
            affected_users_input = dread
            discoverability_input = dread
            reproducibility_input = dread
            damage_potential_input = dread

            dreadc_sim = ctrl.ControlSystemSimulation(control_system_with_cost)
            dread_sim = ctrl.ControlSystemSimulation(control_system_without_cost)

            #Loading input values from brower an rules generated for DREAD
            dread_sim.input['exploitability'] = exploitability_input
            dread_sim.input['affected_users'] = affected_users_input
            dread_sim.input['discoverability'] = discoverability_input
            dread_sim.input['reproducibility'] = reproducibility_input
            dread_sim.input['damage_potential'] = damage_potential_input

            #Compute score fr DREAD
            dread_sim.compute()
            dread_sim_result = dread_sim.output['risk_score']
            
            #Loading input values from brower an rules generated for DREAD-C
            dreadc_sim.input['cost'] = cost_input
            dreadc_sim.input['dread'] = dread_sim_result
            
            #Compute score fr DREAD-C
            dreadc_sim.compute()
            dreadc_sim_result = dreadc_sim.output['risk_score']

            dread_sim_result = round(dread_sim_result, 2)
            dreadc_sim_result = round(dreadc_sim_result, 2)

            decision=""

            if dreadc_sim_result <= 3:
                decision = "Accept Risk."
            elif dreadc_sim_result > 3 and dreadc_sim_result < 4:
                decision = "Transfer Risk."
            elif dreadc_sim_result >= 4 and dreadc_sim_result <= 5:
                decision = "Reduce Risk."
            elif dreadc_sim_result > 5:
                decision = "Avoid Risk."

            print("==================================")
            print(f"COST -- > {cost_input}")
            print(f"EXPLOITABILITY -- > {exploitability_input}")
            print(f"AFFECTED USERS -- > {affected_users_input}")
            print(f"DISCOVERABILITY -- > {discoverability_input}")
            print(f"REPRODUCIBILITY -- > {reproducibility_input}")
            print(f"DAMAGE CONTROL -- > {damage_potential_input}")
            print("------------------------------------")
            print(f"DREAD SCORE -- > {dread_sim_result}")
            print(f"DREAD - C SCORE --> {dreadc_sim_result}")

            obj = {                
                'cost':cost_input,
                'decision': decision,                
                'dread_score': dread_sim_result,
                'dread_c_score': dreadc_sim_result,        
                'exploitability':exploitability_input,
                'affected_users':affected_users_input,
                'discoverability':discoverability_input,
                'reproducibility':reproducibility_input,
                'damage_potential':damage_potential_input,                 
            }

            final_list.append(obj)
        context = {    
            'dread': dread,
            "response": final_list,
            'generate_nums': generate_nums,
            "title": "Incrementing DREAD Parameters with constant COST Parameter",
        }

        
        return render(request, 'core/risk_with_varying_cost_and_constant_dread_parameters.html', context=context)
    
    return render(request, 'core/risk_with_varying_cost_and_constant_dread_parameters.html', {"title": "Incrementing DREAD Parameters with constant COST Parameter", 'generate_nums': 10, 'dread': 10,})

def risk_with_varying_cost_and_constant_dread_parameters_to_excel(request, dread, generate_nums):
    generate_nums = int(generate_nums)
    dread = float(dread)

    numbers = []
    for _ in range(generate_nums):
        number = round(random.uniform(0, 10), 2)
        numbers.append(number)

    numbers.sort()

    final_list = []

    for val in numbers:
        cost_input = val
        exploitability_input = dread
        affected_users_input = dread
        discoverability_input = dread
        reproducibility_input = dread
        damage_potential_input = dread

        dreadc_sim = ctrl.ControlSystemSimulation(control_system_with_cost)
        dread_sim = ctrl.ControlSystemSimulation(control_system_without_cost)

        #Loading input values from brower an rules generated for DREAD
        dread_sim.input['exploitability'] = exploitability_input
        dread_sim.input['affected_users'] = affected_users_input
        dread_sim.input['discoverability'] = discoverability_input
        dread_sim.input['reproducibility'] = reproducibility_input
        dread_sim.input['damage_potential'] = damage_potential_input

        #Compute score fr DREAD
        dread_sim.compute()
        dread_sim_result = dread_sim.output['risk_score']
        
        #Loading input values from brower an rules generated for DREAD-C
        dreadc_sim.input['cost'] = cost_input
        dreadc_sim.input['dread'] = dread_sim_result
        
        #Compute score fr DREAD-C
        dreadc_sim.compute()
        dreadc_sim_result = dreadc_sim.output['risk_score']

        dread_sim_result = round(dread_sim_result, 2)
        dreadc_sim_result = round(dreadc_sim_result, 2)

        decision=""

        if dreadc_sim_result <= 3:
            decision = "Accept Risk."
        elif dreadc_sim_result > 3 and dreadc_sim_result < 4:
            decision = "Transfer Risk."
        elif dreadc_sim_result >= 4 and dreadc_sim_result <= 5:
            decision = "Reduce Risk."
        elif dreadc_sim_result > 5:
            decision = "Avoid Risk."

        print("==================================")
        print(f"COST -- > {cost_input}")
        print(f"EXPLOITABILITY -- > {exploitability_input}")
        print(f"AFFECTED USERS -- > {affected_users_input}")
        print(f"DISCOVERABILITY -- > {discoverability_input}")
        print(f"REPRODUCIBILITY -- > {reproducibility_input}")
        print(f"DAMAGE CONTROL -- > {damage_potential_input}")
        print("------------------------------------")
        print(f"DREAD SCORE -- > {dread_sim_result}")
        print(f"DREAD - C SCORE --> {dreadc_sim_result}")

        obj = {
            'cost':cost_input,
            'decision': decision,            
            'exploitability':exploitability_input,
            'affected_users':affected_users_input,
            'discoverability':discoverability_input,
            'reproducibility':reproducibility_input,
            'damage_potential':damage_potential_input,
            'dread_score': dread_sim_result,
            'dread_c_score': dreadc_sim_result,
        }

        final_list.append(obj)

        filename = f"varying_cost_{slugify(datetime.now().isoformat())}.xls"

        response = HttpResponse(content_type="application/ms-excel")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb = xlwt.Workbook(encoding="utf-8")
        ws = wb.add_sheet("DREAD-C")

        # Heading styles
        heading_font_style = xlwt.XFStyle()
        heading_font_style.font.bold = True
        heading_font_style.font.underline = True

        # Sub Heading styles
        sub_heading_font_style = xlwt.XFStyle()
        sub_heading_font_style.font.bold = True
        sub_heading_font_style.font.italic = True

        # Column Heading styles
        column_heading_font_style = xlwt.XFStyle()
        column_heading_font_style.font.bold = True

        # Adding Equity Bank Title
        ws.write(0, 0, "DREAD-C Risk Calculation", heading_font_style) 

        # Sheet body, remaining rows
        font_style = xlwt.XFStyle()
        # Writing column headings
        ws.write(2, 0, "#", sub_heading_font_style)
        ws.write(2, 1, "Exploitability", sub_heading_font_style)
        ws.write(2, 2, "Affected Users", sub_heading_font_style)
        ws.write(2, 3, "Discoverability", sub_heading_font_style)
        ws.write(2, 4, "Reproducibility", sub_heading_font_style)
        ws.write(2, 5, "Damage Potential", sub_heading_font_style)
        ws.write(2, 6, "Cost", sub_heading_font_style)
        ws.write(2, 7, "DREAD Score", sub_heading_font_style)
        ws.write(2, 8, "DREAD-C Score", sub_heading_font_style)
        ws.write(2, 9, "Decision", sub_heading_font_style)

        row_num = 3
        for obj in final_list:
            ws.write(row_num, 0, row_num - 2, font_style)  # Serial number
            ws.write(row_num, 1, obj["exploitability"], font_style)
            ws.write(row_num, 2, obj["affected_users"], font_style)
            ws.write(row_num, 3, obj["discoverability"], font_style)
            ws.write(row_num, 4, obj["reproducibility"], font_style)
            ws.write(row_num, 5, obj["damage_potential"], font_style)
            ws.write(row_num, 6, obj["cost"], font_style)
            ws.write(row_num, 7, obj["dread_score"], font_style)
            ws.write(row_num, 8, obj["dread_c_score"], font_style)
            ws.write(row_num, 9, obj["decision"], font_style)
            row_num += 1
        # for current_list in final_list:
        #     ws.write(row_num, 0, obj["exploitability"], font_style)
        #     for obj in current_list:                
        #         ws.write(row_num, 1, obj["exploitability"], font_style)
        #         ws.write(row_num, 2, obj["affected_users"], font_style)
        #         ws.write(row_num, 3, obj["discoverability"], font_style)
        #         ws.write(row_num, 4, obj["reproducibility"], font_style)
        #         ws.write(row_num, 5, obj["damage_potential"], font_style)
        #         ws.write(row_num, 6, obj["cost"], font_style)
        #         ws.write(row_num, 7, obj["dread_score"], font_style)
        #         ws.write(row_num, 8, obj["dread_c_score"], font_style)
        #         ws.write(row_num, 9, obj["decision"], font_style)
        #     row_num = row_num + 1

        wb.save(response)
        return response


def risk_with_varying_dread_and_constant_cost_parameters_to_excel(request):
    generate_nums = int(request.POST.get('generate_nums', 1))
    dread = float(request.POST.get('dread', 10))

    numbers = []
    for _ in range(generate_nums):
        number = round(random.uniform(0, 10), 2)
        numbers.append(number)

    numbers.sort()

    final_list = []

    for val in numbers:
        cost_input = val
        exploitability_input = dread
        affected_users_input = dread
        discoverability_input = dread
        reproducibility_input = dread
        damage_potential_input = dread

        dreadc_sim = ctrl.ControlSystemSimulation(control_system_with_cost)
        dread_sim = ctrl.ControlSystemSimulation(control_system_without_cost)

        #Loading input values from brower an rules generated for DREAD
        dread_sim.input['exploitability'] = exploitability_input
        dread_sim.input['affected_users'] = affected_users_input
        dread_sim.input['discoverability'] = discoverability_input
        dread_sim.input['reproducibility'] = reproducibility_input
        dread_sim.input['damage_potential'] = damage_potential_input

        #Compute score fr DREAD
        dread_sim.compute()
        dread_sim_result = dread_sim.output['risk_score']
        
        #Loading input values from brower an rules generated for DREAD-C
        dreadc_sim.input['cost'] = cost_input
        dreadc_sim.input['dread'] = dread_sim_result
        
        #Compute score fr DREAD-C
        dreadc_sim.compute()
        dreadc_sim_result = dreadc_sim.output['risk_score']

        dread_sim_result = round(dread_sim_result, 2)
        dreadc_sim_result = round(dreadc_sim_result, 2)

        decision=""

        if dreadc_sim_result <= 3:
            decision = "Accept Risk."
        elif dreadc_sim_result > 3 and dreadc_sim_result < 4:
            decision = "Transfer Risk."
        elif dreadc_sim_result >= 4 and dreadc_sim_result <= 5:
            decision = "Reduce Risk."
        elif dreadc_sim_result > 5:
            decision = "Avoid Risk."

        print("==================================")
        print(f"COST -- > {cost_input}")
        print(f"EXPLOITABILITY -- > {exploitability_input}")
        print(f"AFFECTED USERS -- > {affected_users_input}")
        print(f"DISCOVERABILITY -- > {discoverability_input}")
        print(f"REPRODUCIBILITY -- > {reproducibility_input}")
        print(f"DAMAGE CONTROL -- > {damage_potential_input}")
        print("------------------------------------")
        print(f"DREAD SCORE -- > {dread_sim_result}")
        print(f"DREAD - C SCORE --> {dreadc_sim_result}")

        obj = {
            'cost':cost_input,
            'decision': decision,            
            'exploitability':exploitability_input,
            'affected_users':affected_users_input,
            'discoverability':discoverability_input,
            'reproducibility':reproducibility_input,
            'damage_potential':damage_potential_input,
            'dread_score': dread_sim_result,
            'dread_c_score': dreadc_sim_result,
        }

        final_list.append(obj)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "DREAD Risk Results"

        # Write header
        headers = list(final_list[0].keys())
        ws.append(headers)

        # Write data rows
        for item in final_list:
            ws.append(list(item.values()))

        # Auto-adjust column widths
        for col_num, _ in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = 18

        # Create HTTP response
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"risk_results_dreadc.xlsx"
        response['Content-Disposition'] = f'attachment; filename={filename}'

        wb.save(response)
        return response
